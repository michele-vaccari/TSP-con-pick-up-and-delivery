from tsppd.problem.generator import Generator
from tsppd.problem.instance import Instance
from tsppd.solver.bruteForceEnumerator import BruteForceEnumerator
from tsppd.solver.oneilHoffmanEnumerator import OneilHoffmanEnumerator
from tsppd.solver.greedyTemplate import GreedyTemplate
from tsppd.solver.greedyPickupFirst import GreedyPickupFirst
from tsppd.solver.greedyRequestOrder import GreedyRequestOrder
from tsppd.solver.greedyNearestNeighbor import GreedyNearestNeighbor
from tsppd.solver.greedyRandom import GreedyRandom
from tsppd.solver.cityInsert import CityInsert
from tsppd.solver.citySwap import CitySwap
from tsppd.solver.simulatedAnnealing import SimulatedAnnealing
from tsppd.solver.tabuSearch import TabuSearch
from tsppd.solver.largeNeighborhoodSearch import LargeNeighborhoodSearch
from tsppd.solver.multiStartLocalSearch import MultiStartLocalSearch
from tsppd.solver.greedyRandomizedAdaptiveSearchProcedure import GreedyRandomizedAdaptiveSearchProcedure
from tsppd.solver.geneticAlgorithm import GeneticAlgoritm

from tsppd.benchmark.bruteForceEnumeratorBenchmark import BruteForceEnumeratorBenchmark
from tsppd.benchmark.oneilHoffmanEnumeratorBenchmark import OneilHoffmanEnumeratorBenchmark
from tsppd.benchmark.greedyPickupFirstBenchmark import GreedyPickupFirstBenchmark
from tsppd.benchmark.greedyRequestOrderBenchmark import GreedyRequestOrderBenchmark
from tsppd.benchmark.greedyNearestNeighborBenchmark import GreedyNearestNeighborBenchmark
from tsppd.benchmark.greedyRandomBenchmark import GreedyRandomBenchmark
from tsppd.benchmark.cityInsertBenchmark import CityInsertBenchmark
from tsppd.benchmark.citySwapBenchmark import CitySwapBenchmark

import click
import openpyxl
import os
from stopwatch import Stopwatch

@click.group()
def cli():
    pass

@click.command()
@click.option('--requests', help='Number of requests.', type = click.INT)
@click.option('--weights-as-euclidean-distance/--weights-random', default=False, help='Weights are the Euclidean distance of the nodes / Weights are randomly generated.')
@click.option('--output-instance-path', default=None, help='Path where to save the instance in JSON format.', type = click.STRING)
def generate_instance(requests, weights_as_euclidean_distance, output_instance_path):
    """Generates an instance of the tsppd problem."""
    generator = Generator(requests, weights_as_euclidean_distance)
    instance = generator.generate_instance()
    print(instance)

    if output_instance_path == None:
        return
    
    instance.save_to_json(output_instance_path)

cli.add_command(generate_instance)

@click.command()
@click.option('--input-instance-path', required=True, default=None, help='Path where read the instance in JSON format.', type = click.STRING)
@click.option('--solver-method', required=True, default=None, help='Choose the solver method to use.', type=click.Choice(['brute-force-enumerator', 'oneil-hoffman-enumerator', 'greedy-pickup-first', 'greedy-request-order', 'greedy-nearest-neighbor', 'greedy-random', 'city-swap', 'city-insert', 'simulated-annealing', 'tabu-search', 'large-neighborhood-search', 'multi-start-local-search', 'greedy-randomized-adaptive-search-procedure', 'greedy-randomized-adaptive-search-procedure-with-path-relinking', 'genetic-algorithm'], case_sensitive=False))
def solve(input_instance_path, solver_method):
    """Solve an instance of the tsppd problem."""
    print("\n\nREADED INSTANCE\n\n")
    instance = Instance.read_from_json(input_instance_path)
    print(instance)

    if (solver_method == "brute-force-enumerator"):
        enumerator = BruteForceEnumerator(instance)
        solution = enumerator.calculate_solution()
        print("\nBEST SOLUTION: \n")
        print(solution)
    elif (solver_method == "oneil-hoffman-enumerator"):
        enumerator = OneilHoffmanEnumerator(instance)
        solution = enumerator.calculate_solution()
        print("\nBEST SOLUTION: \n")
        print(solution)
    elif (solver_method == "greedy-pickup-first"):
        greedyPickupFirst = GreedyPickupFirst()
        greedyTemplate = GreedyTemplate(instance, greedyPickupFirst)
        solution = greedyTemplate.calculate_solution()
        print("\nGREEDY PICKUP-FIRST SOLUTION: \n")
        print(solution)
    elif (solver_method == "greedy-request-order"):
        greedyRequestOrder = GreedyRequestOrder()
        greedyTemplate = GreedyTemplate(instance, greedyRequestOrder)
        solution = greedyTemplate.calculate_solution()
        print("\nGREEDY REQUEST-ORDER SOLUTION: \n")
        print(solution)
    elif (solver_method == "greedy-nearest-neighbor"):
        greedyNearestNeighbor = GreedyNearestNeighbor()
        greedyTemplate = GreedyTemplate(instance, greedyNearestNeighbor)
        solution = greedyTemplate.calculate_solution()
        print("\nGREEDY NEAREST-NEIGHBOR SOLUTION: \n")
        print(solution)
    elif (solver_method == "greedy-random"):
        greedyRandom = GreedyRandom()
        greedyTemplate = GreedyTemplate(instance, greedyRandom)
        solution = greedyTemplate.calculate_solution()
        print("\nGREEDY RANDOM SOLUTION: \n")
        print(solution)
    elif (solver_method == "city-swap"):
        best_greedy_solution = compute_all_greedy_and_get_best_greedy_solution(instance)

        citySwap = CitySwap(instance, best_greedy_solution)
        citySwapSolution = citySwap.calculate_solution()

        print("\nCITY SWAP SOLUTION: \n")
        print(citySwapSolution)
    elif (solver_method == "city-insert"):
        best_greedy_solution = compute_all_greedy_and_get_best_greedy_solution(instance)

        cityInsert = CityInsert(instance, best_greedy_solution)
        cityInsertSolution = cityInsert.calculate_solution()

        print("\nCITY INSERT SOLUTION: \n")
        print(cityInsertSolution)
    elif (solver_method == "simulated-annealing"):
        best_greedy_solution = compute_all_greedy_and_get_best_greedy_solution(instance)

        tabuSearch = SimulatedAnnealing(instance, best_greedy_solution)
        tabuSearchSolution = tabuSearch.calculate_solution()

        print("\nSIMULATED ANNEALING SOLUTION: \n")
        print(tabuSearchSolution)
    elif (solver_method == "tabu-search"):
        best_greedy_solution = compute_all_greedy_and_get_best_greedy_solution(instance)

        tabuSearch = TabuSearch(instance, best_greedy_solution)
        tabuSearchSolution = tabuSearch.calculate_solution()

        print("\nTABU SEARCH SOLUTION: \n")
        print(tabuSearchSolution)
    elif (solver_method == "large-neighborhood-search"):
        best_greedy_solution = compute_all_greedy_and_get_best_greedy_solution(instance)

        largeNeighborhoodSearch = LargeNeighborhoodSearch(instance, best_greedy_solution)
        largeNeighborhoodSearchSolution = largeNeighborhoodSearch.calculate_solution()

        print("\nLARGE NEIGHBORHOOD SEARCH SOLUTION: \n")
        print(largeNeighborhoodSearchSolution)
    elif (solver_method == "multi-start-local-search"):
        multiStartLocalSearch = MultiStartLocalSearch(instance)
        multiStartLocalSearchSolution = multiStartLocalSearch.calculate_solution()

        print("\nMULTI START LOCAL SEARCH SOLUTION: \n")
        print(multiStartLocalSearchSolution)
    elif (solver_method == "greedy-randomized-adaptive-search-procedure"):
        greedyRandomizedAdaptiveSearchProcedure = GreedyRandomizedAdaptiveSearchProcedure(instance)
        greedyRandomizedAdaptiveSearchProcedureSolution = greedyRandomizedAdaptiveSearchProcedure.calculate_solution(False)

        print("\nGREEDY RANDOMIZED ADAPTIVE SEARCH PROCEDURE SOLUTION: \n")
        print(greedyRandomizedAdaptiveSearchProcedureSolution)
    elif (solver_method == "greedy-randomized-adaptive-search-procedure-with-path-relinking"):
        greedyRandomizedAdaptiveSearchProcedure = GreedyRandomizedAdaptiveSearchProcedure(instance)
        greedyRandomizedAdaptiveSearchProcedureSolution = greedyRandomizedAdaptiveSearchProcedure.calculate_solution(True)

        print("\nGREEDY RANDOMIZED ADAPTIVE SEARCH PROCEDURE WITH PATH RELINKING SOLUTION: \n")
        print(greedyRandomizedAdaptiveSearchProcedureSolution)
    elif (solver_method == "genetic-algorithm"):
        geneticAlgorithm = GeneticAlgoritm(instance)
        geneticAlgorithmSolution = geneticAlgorithm.calculate_solution()

        print("\nGENETIC ALGORITHM SOLUTION: \n")
        print(geneticAlgorithmSolution)

cli.add_command(solve)

def compute_all_greedy_and_get_best_greedy_solution(instance: Instance):
    solutions = dict()

    greedyPickupFirst = GreedyPickupFirst()
    greedyTemplate = GreedyTemplate(instance, greedyPickupFirst)
    greedyPickupFirstSolution = greedyTemplate.calculate_solution()
    solutions["Greedy Pickup First"] = greedyPickupFirstSolution

    greedyRequestOrder = GreedyRequestOrder()
    greedyTemplate = GreedyTemplate(instance, greedyRequestOrder)
    greedyRequestOrderSolution = greedyTemplate.calculate_solution()
    solutions["Greedy Request Order"] = greedyRequestOrderSolution

    greedyNearestNeighbor = GreedyNearestNeighbor()
    greedyTemplate = GreedyTemplate(instance, greedyNearestNeighbor)
    greedyNearestNeighborSolution = greedyTemplate.calculate_solution()
    solutions["Greedy Nearest Neighbor"] = greedyNearestNeighborSolution

    greedyRandom = GreedyRandom()
    greedyTemplate = GreedyTemplate(instance, greedyRandom)
    greedyRandomSolution = greedyTemplate.calculate_solution()
    solutions["Greedy Random"] = greedyRandomSolution

    best_solution = min(solutions.values())
    best_solution_method = ([k for k,v in solutions.items() if v == best_solution])
    print("The best greedy solution is obtain with: " + best_solution_method[0])
    print(best_solution)
    return best_solution

def benchmark_instance_generator(requests_benchmark_start, requests_benchmark_end, workbook_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Genero istanze da R = 2..30
    # asimmetriche
    weights_as_euclidean_distance = False

    # title
    row_index = 1
    column_index = 1
    sheet.cell(row = row_index, column = column_index).value = "Generazione delle istanze da R = 2 a R = 30"

    # header
    row_index += 1
    sheet.cell(row=row_index , column = column_index).value = "Numero di richieste"
    column_index += 1
    sheet.cell(row=row_index , column = column_index).value = "Durata (in secondi)"

    # data
    for requests in range(requests_benchmark_start,requests_benchmark_end + 1):
        row_index += 1
        sheet.cell(row=row_index, column=1).value = requests
        stopwatch = Stopwatch(2)
        generator = Generator(requests, weights_as_euclidean_distance)
        instance = generator.generate_instance()
        output_instance_path = "benchmark/{}-request-weights-random.json".format(requests)
        instance.save_to_json(output_instance_path)
        stopwatch.stop()
        sheet.cell(row=row_index, column=2).value = stopwatch.duration
        
    workbook.save(workbook_path)

@click.command()
@click.option('--output-excel-spreadsheets-dir-path', default=None, help='Directory path where to save the excel spreadsheets.', type = click.STRING)
def benchmark(output_excel_spreadsheets_dir_path):
    """Generates benchmarks of the algorithms implemented to solve the tsppd problem."""

    requests_benchmark_start = 2
    requests_benchmark_end = 30

    # Instance generator
    algoritm_name = "Instance generator"
    print("{} benchmark - START".format(algoritm_name))
    stopwatch = Stopwatch(2)
    workbook_path = os.path.join(output_excel_spreadsheets_dir_path, "{} benchmark.xlsx".format(algoritm_name))
    benchmark_instance_generator(requests_benchmark_start, requests_benchmark_end, workbook_path)
    stopwatch.stop()
    print("{} benchmark - END".format(algoritm_name))
    print("See the results on: {}".format(workbook_path))
    print("Time taken for {} benchmark: {} seconds\n".format(algoritm_name.lower(), stopwatch.duration))

    enumerator_requests_benchmark_start = 2
    enumerator_requests_benchmark_end = 3

    # Brute force enumerator
    algoritm_name = "Brute force enumerator"
    print("{} benchmark - START".format(algoritm_name))
    workbook_path = os.path.join(output_excel_spreadsheets_dir_path, "{} benchmark.xlsx".format(algoritm_name))
    stopwatch.restart()
    bruteForceEnumeratorBenchmark = BruteForceEnumeratorBenchmark(enumerator_requests_benchmark_start, enumerator_requests_benchmark_end, workbook_path)
    bruteForceEnumeratorBenchmark.benchmark()
    stopwatch.stop()
    print("{} benchmark - END".format(algoritm_name))
    print("See the results on: {}".format(workbook_path))
    print("Time taken for {} benchmark: {} seconds\n".format(algoritm_name.lower(), stopwatch.duration))

    # Oneil Hoffman enumerator
    algoritm_name = "Oneil Hoffman enumerator"
    print("{} benchmark - START".format(algoritm_name))
    workbook_path = os.path.join(output_excel_spreadsheets_dir_path, "{} benchmark.xlsx".format(algoritm_name))
    stopwatch.restart()
    oneilHoffmanEnumeratorBenchmark = OneilHoffmanEnumeratorBenchmark(enumerator_requests_benchmark_start, enumerator_requests_benchmark_end, workbook_path)
    oneilHoffmanEnumeratorBenchmark.benchmark()
    stopwatch.stop()
    print("{} benchmark - END".format(algoritm_name))
    print("See the results on: {}".format(workbook_path))
    print("Time taken for {} benchmark: {} seconds\n".format(algoritm_name.lower(), stopwatch.duration))

    greedy_requests_benchmark_start = 2
    greedy_requests_benchmark_end = 3

    # Greedy Pickup First
    algoritm_name = "Greedy Pickup First"
    print("{} benchmark - START".format(algoritm_name))
    workbook_path = os.path.join(output_excel_spreadsheets_dir_path, "{} benchmark.xlsx".format(algoritm_name))
    stopwatch.restart()
    greedyPickupFirstBenchmark = GreedyPickupFirstBenchmark(greedy_requests_benchmark_start, greedy_requests_benchmark_end, workbook_path)
    greedyPickupFirstBenchmark.benchmark()
    stopwatch.stop()
    print("{} benchmark - END".format(algoritm_name))
    print("See the results on: {}".format(workbook_path))
    print("Time taken for {} benchmark: {} seconds\n".format(algoritm_name.lower(), stopwatch.duration))

    # Greedy Request Order
    algoritm_name = "Greedy Request Order"
    print("{} benchmark - START".format(algoritm_name))
    workbook_path = os.path.join(output_excel_spreadsheets_dir_path, "{} benchmark.xlsx".format(algoritm_name))
    stopwatch.restart()
    greedyRequestOrderBenchmark = GreedyRequestOrderBenchmark(greedy_requests_benchmark_start, greedy_requests_benchmark_end, workbook_path)
    greedyRequestOrderBenchmark.benchmark()
    stopwatch.stop()
    print("{} benchmark - END".format(algoritm_name))
    print("See the results on: {}".format(workbook_path))
    print("Time taken for {} benchmark: {} seconds\n".format(algoritm_name.lower(), stopwatch.duration))

    # Greedy Nearest Neighbor
    algoritm_name = "Greedy Nearest Neighbor"
    print("{} benchmark - START".format(algoritm_name))
    workbook_path = os.path.join(output_excel_spreadsheets_dir_path, "{} benchmark.xlsx".format(algoritm_name))
    stopwatch.restart()
    greedyNearestNeighborBenchmark = GreedyNearestNeighborBenchmark(greedy_requests_benchmark_start, greedy_requests_benchmark_end, workbook_path)
    greedyNearestNeighborBenchmark.benchmark()
    stopwatch.stop()
    print("{} benchmark - END".format(algoritm_name))
    print("See the results on: {}".format(workbook_path))
    print("Time taken for {} benchmark: {} seconds\n".format(algoritm_name.lower(), stopwatch.duration))

    # Greedy Random
    algoritm_name = "Greedy Random"
    print("{} benchmark - START".format(algoritm_name))
    workbook_path = os.path.join(output_excel_spreadsheets_dir_path, "{} benchmark.xlsx".format(algoritm_name))
    stopwatch.restart()
    greedyRandomBenchmark = GreedyRandomBenchmark(greedy_requests_benchmark_start, greedy_requests_benchmark_end, workbook_path)
    greedyRandomBenchmark.benchmark()
    stopwatch.stop()
    print("{} benchmark - END".format(algoritm_name))
    print("See the results on: {}".format(workbook_path))
    print("Time taken for {} benchmark: {} seconds\n".format(algoritm_name.lower(), stopwatch.duration))

    # Compute best greedy solution
    greedy_solutions = dict()
    for key, value in greedyPickupFirstBenchmark.solutions.items():
        if key in greedy_solutions:
            greedy_solutions[key].append(["Greedy Pickup First", value])
        else:
            greedy_solutions[key] = [["Greedy Pickup First", value]]
    for key, value in greedyRequestOrderBenchmark.solutions.items():
        if key in greedy_solutions:
            greedy_solutions[key].append(["Greedy Request Order", value])
        else:
            greedy_solutions[key] = [["Greedy Request Order", value]]
    for key, value in greedyNearestNeighborBenchmark.solutions.items():
        if key in greedy_solutions:
            greedy_solutions[key].append(["Greedy Nearest Neighbor", value])
        else:
            greedy_solutions[key] = [["Greedy Nearest Neighbor", value]]
    for key, value in greedyRandomBenchmark.solutions.items():
        if key in greedy_solutions:
            greedy_solutions[key].append(["Greedy Random", value])
        else:
            greedy_solutions[key] = [["Greedy Random", value]]

    print("Greedy best solutions:")
    greedy_best_solutions = dict()
    for key, value in greedy_solutions.items():
        greedy_request_solution = dict()
        for element in value:
            greedy_request_solution[element[0]] = element[1]
        best_greedy_solution = min(greedy_request_solution.values())
        best_solution_method = ([k for k,v in greedy_request_solution.items() if v == best_greedy_solution])
        greedy_best_solutions[key] = [best_solution_method, best_greedy_solution]
        print("* for R = {} the best solution is obtained with {} with solution path: {} and cost {}".format(key, best_solution_method[0], ' -> '.join(str(node_id) for node_id in best_greedy_solution.solution_nodes_id), best_greedy_solution.cost))

    # City Insert
    algoritm_name = "City Insert"
    print("\n{} benchmark - START".format(algoritm_name))
    workbook_path = os.path.join(output_excel_spreadsheets_dir_path, "{} benchmark.xlsx".format(algoritm_name))
    stopwatch.restart()
    cityInsertBenchmark = CityInsertBenchmark(greedy_best_solutions, greedy_requests_benchmark_start, greedy_requests_benchmark_end, workbook_path)
    cityInsertBenchmark.benchmark()
    stopwatch.stop()
    print("{} benchmark - END".format(algoritm_name))
    print("See the results on: {}".format(workbook_path))
    print("Time taken for {} benchmark: {} seconds\n".format(algoritm_name.lower(), stopwatch.duration))

    # City Swap
    algoritm_name = "City Swap"
    print("\n{} benchmark - START".format(algoritm_name))
    workbook_path = os.path.join(output_excel_spreadsheets_dir_path, "{} benchmark.xlsx".format(algoritm_name))
    stopwatch.restart()
    citySwapBenchmark = CitySwapBenchmark(greedy_best_solutions, greedy_requests_benchmark_start, greedy_requests_benchmark_end, workbook_path)
    citySwapBenchmark.benchmark()
    stopwatch.stop()
    print("{} benchmark - END".format(algoritm_name))
    print("See the results on: {}".format(workbook_path))
    print("Time taken for {} benchmark: {} seconds\n".format(algoritm_name.lower(), stopwatch.duration))

cli.add_command(benchmark)

if __name__ == '__main__':

    cli()