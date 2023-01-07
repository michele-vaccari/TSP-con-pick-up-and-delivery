from ..utils.observerPattern import Subject, Observer

from tsppd.problem.instance import Instance
from tsppd.solver.citySwap import CitySwap

import openpyxl
from stopwatch import Stopwatch

class CitySwapBenchmark(Observer):
    def __init__(self, greedy_best_solutions, requests_benchmark_start , requests_benchmark_end, workbook_path):
        self._greedy_best_solutions = greedy_best_solutions
        self._requests_benchmark_start = requests_benchmark_start
        self._requests_benchmark_end = requests_benchmark_end
        self._workbook_path = workbook_path

        self._update_row_index = None
        self._stopwatch = None
        self._current_requests = None

        self._solution_trend_sheet = None
        self._solution_trend_sheet_row_index = None
        self._best_solution_trend_sheet_row_index = None

    def new_solution_found(self, subject: Subject) -> None:
        self._stopwatch.stop()

        self._solution_trend_sheet_row_index += 1
        self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=1).value = subject.feasible_solutions_counter
        self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=2).value = self._stopwatch.duration
        self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=3).value = ' -> '.join(str(node_id) for node_id in subject.current_solution.solution_nodes_id)
        self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=4).value = subject.current_solution.cost

        self._stopwatch.start()
    
    def new_best_solution_found(self, subject: Subject) -> None:
        self._stopwatch.stop()

        self._best_solution_trend_sheet_row_index += 1
        self._best_solution_trend_sheet.cell(row=self._best_solution_trend_sheet_row_index, column=1).value = subject.feasible_solutions_counter
        self._best_solution_trend_sheet.cell(row=self._best_solution_trend_sheet_row_index, column=2).value = self._stopwatch.duration
        self._best_solution_trend_sheet.cell(row=self._best_solution_trend_sheet_row_index, column=3).value = ' -> '.join(str(node_id) for node_id in subject.best_solution.solution_nodes_id)
        self._best_solution_trend_sheet.cell(row=self._best_solution_trend_sheet_row_index, column=4).value = subject.best_solution.cost

        self._stopwatch.start()
    
    def benchmark(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Riassunto"

        # title of resume sheet
        row_index = 1
        column_index = 1
        sheet.cell(row = row_index, column = column_index).value = "Performance nel tempo dell'algoritmo"

        # header of resume sheet
        row_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Numero di richieste"
        column_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Durata (in secondi)"
        column_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Valore della migliore soluzione"
        column_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Costo della migliore soluzione"
        column_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Soluzioni ammissibili esplorate"

        # data
        for requests in range(self._requests_benchmark_start, self._requests_benchmark_end + 1):
            print("* computation with R = {} ...".format(requests))
            self._current_requests = requests

            # header of solution trend sheet
            self._solution_trend_sheet = workbook.create_sheet("Trend soluzioni R = {}".format(self._current_requests))
            self._solution_trend_sheet_row_index = 1
            self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=1).value = "Numero iterazione"
            self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=2).value = "Nuova soluzione scoperta al tempo (s)"
            self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=3).value = "Nodi della soluzione"
            self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=4).value = "Costo della soluzione"

            # header of best solution trend sheet
            self._best_solution_trend_sheet = workbook.create_sheet("Trend migliori soluzioni R = {}".format(self._current_requests))
            self._best_solution_trend_sheet_row_index = 1
            self._best_solution_trend_sheet.cell(row=self._best_solution_trend_sheet_row_index, column=1).value = "Numero iterazione"
            self._best_solution_trend_sheet.cell(row=self._best_solution_trend_sheet_row_index, column=2).value = "Nuova migliore soluzione scoperta al tempo (s)"
            self._best_solution_trend_sheet.cell(row=self._best_solution_trend_sheet_row_index, column=3).value = "Nodi della soluzione"
            self._best_solution_trend_sheet.cell(row=self._best_solution_trend_sheet_row_index, column=4).value = "Costo della soluzione"

            self._stopwatch = Stopwatch(2)
        
            input_instance_path = "benchmark/{}-request-weights-random.json".format(requests)
            instance = Instance.read_from_json(input_instance_path)

            citySwap = CitySwap(instance, self._greedy_best_solutions[requests][1])
            citySwap.attach(self)
            solution = citySwap.calculate_solution()
            citySwap.detach(self)

            self._stopwatch.stop()

            # value of resume sheet
            row_index += 1
            sheet.cell(row=row_index, column=1).value = requests
            sheet.cell(row=row_index, column=2).value = self._stopwatch.duration
            sheet.cell(row=row_index, column=3).value = ' -> '.join(str(node_id) for node_id in solution.solution_nodes_id)
            sheet.cell(row=row_index, column=4).value = solution.cost
            sheet.cell(row=row_index, column=5).value = citySwap.feasible_solutions_counter
            
        workbook.save(self._workbook_path)