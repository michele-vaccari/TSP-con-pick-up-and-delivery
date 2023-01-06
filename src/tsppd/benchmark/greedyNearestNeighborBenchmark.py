from tsppd.problem.instance import Instance
from tsppd.solver.greedyTemplate import GreedyTemplate
from tsppd.solver.greedyNearestNeighbor import GreedyNearestNeighbor

import openpyxl
from stopwatch import Stopwatch

class GreedyNearestNeighborBenchmark():
    def __init__(self, requests_benchmark_start , requests_benchmark_end, workbook_path):
        self._requests_benchmark_start = requests_benchmark_start
        self._requests_benchmark_end = requests_benchmark_end
        self._workbook_path = workbook_path

        self._stopwatch = None
        self._solutions = dict()
    
    @property
    def solutions(self):
        return self._solutions
    
    def benchmark(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Riassunto"

        # title of resume sheet
        row_index = 1
        column_index = 1
        sheet.cell(row = row_index, column = column_index).value = "Performance nel tempo dell'algoritmo"

        # header of resume sheet
        row_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Numero di richieste"
        column_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Durata (in secondi)"
        column_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Valore della soluzione"
        column_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Costo della soluzione"

        # data
        for requests in range(self._requests_benchmark_start, self._requests_benchmark_end + 1):
            self._stopwatch = Stopwatch(2)
        
            input_instance_path = "benchmark/{}-request-weights-random.json".format(requests)
            instance = Instance.read_from_json(input_instance_path)

            greedyNearestNeighbor = GreedyNearestNeighbor()
            greedyTemplate = GreedyTemplate(instance, greedyNearestNeighbor)
            solution = greedyTemplate.calculate_solution()

            self._stopwatch.stop()
            self._solutions[requests] = solution

            # value of resume sheet
            row_index += 1
            sheet.cell(row=row_index, column=1).value = requests
            sheet.cell(row=row_index, column=2).value = self._stopwatch.duration
            sheet.cell(row=row_index, column=3).value = ' -> '.join(str(node_id) for node_id in solution.solution_nodes_id)
            sheet.cell(row=row_index, column=4).value = solution.cost
            
        workbook.save(self._workbook_path)