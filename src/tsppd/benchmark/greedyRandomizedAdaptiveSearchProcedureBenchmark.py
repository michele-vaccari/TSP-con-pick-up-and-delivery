from ..utils.observerPattern import Subject, Observer

from tsppd.problem.instance import Instance
from tsppd.solver.greedyRandomizedAdaptiveSearchProcedure import GreedyRandomizedAdaptiveSearchProcedure

import openpyxl
from stopwatch import Stopwatch

class GreedyRandomizedAdaptiveSearchProcedureBenchmark(Observer):
    def __init__(self, requests_benchmark_start , requests_benchmark_end, workbook_path):
        self._requests_benchmark_start = requests_benchmark_start
        self._requests_benchmark_end = requests_benchmark_end
        self._workbook_path = workbook_path

        self._update_row_index = None
        self._stopwatch = None
        self._current_requests = None

        self._solution_trend_sheet = None
        self._solution_trend_sheet_row_index = None

    def new_solution_found(self, subject: Subject) -> None:
        self._stopwatch.stop()

        self._solution_trend_sheet_row_index += 1
        self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=1).value = subject.feasible_solutions_counter
        self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=2).value = self._stopwatch.duration
        self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=3).value = ' -> '.join(str(node_id) for node_id in subject.current_solution.solution_nodes_id)
        self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=4).value = subject.current_solution.cost

        self._stopwatch.start()
    
    def new_best_solution_found(self, subject: Subject) -> None:
        return NotImplemented
    
    def benchmark(self, path_relinking = False):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Riassunto"

        # title of resume sheet
        row_index = 1
        column_index = 1
        sheet.cell(row = row_index, column = column_index).value = "Performance nel tempo dell'algoritmo"

        # header of resume sheet
        row_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Numero di richieste"
        column_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Durata (in secondi)"
        column_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Valore della migliore soluzione"
        column_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Costo della migliore soluzione"
        column_index += 1
        sheet.cell(row=row_index , column = column_index).value = "Soluzioni ammissibili esplorate"
        column_index += 1
        sheet.cell(row=row_index, column = column_index).value = "Numero massimo di iterazioni"
        column_index += 1
        sheet.cell(row=row_index, column = column_index).value = "Dimensione della restricted candidate list"

        # data
        for requests in range(self._requests_benchmark_start, self._requests_benchmark_end + 1):
            self._current_requests = requests

            # header of solution trend sheet
            self._solution_trend_sheet = workbook.create_sheet("Trend soluzioni R = {}".format(self._current_requests))
            self._solution_trend_sheet_row_index = 1
            self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=1).value = "Numero iterazione"
            self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=2).value = "Nuova soluzione scoperta al tempo (s)"
            self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=3).value = "Nodi della soluzione"
            self._solution_trend_sheet.cell(row=self._solution_trend_sheet_row_index, column=4).value = "Costo della soluzione"

            self._stopwatch = Stopwatch(2)
        
            input_instance_path = "benchmark/{}-request-weights-random.json".format(requests)
            instance = Instance.read_from_json(input_instance_path)

            greedyRandomizedAdaptiveSearchProcedure = GreedyRandomizedAdaptiveSearchProcedure(instance)
            greedyRandomizedAdaptiveSearchProcedure.attach(self)
            solution = greedyRandomizedAdaptiveSearchProcedure.calculate_solution(path_relinking)
            greedyRandomizedAdaptiveSearchProcedure.detach(self)

            self._stopwatch.stop()

            # value of resume sheet
            row_index += 1
            sheet.cell(row=row_index, column=1).value = requests
            sheet.cell(row=row_index, column=2).value = self._stopwatch.duration
            sheet.cell(row=row_index, column=3).value = ' -> '.join(str(node_id) for node_id in solution.solution_nodes_id)
            sheet.cell(row=row_index, column=4).value = solution.cost
            sheet.cell(row=row_index, column=5).value = greedyRandomizedAdaptiveSearchProcedure.feasible_solutions_counter
            sheet.cell(row=row_index, column=6).value = greedyRandomizedAdaptiveSearchProcedure.max_iterations_number
            sheet.cell(row=row_index, column=7).value = greedyRandomizedAdaptiveSearchProcedure._restricted_candidate_list_items
            
        workbook.save(self._workbook_path)